import openpyxl
# Give the location of the file
def getCoulums(filename):
    My_wb_obj = openpyxl.load_workbook(filename)
    my_sheet_obj = My_wb_obj.active
    return my_sheet_obj.max_column   
def getRows(filename):
    My_wb_obj = openpyxl.load_workbook(filename)
    my_sheet_obj = My_wb_obj.active
    return my_sheet_obj.max_row   
def readCellValue(Row,column,path):
    wb_obj = openpyxl.load_workbook(path)
    my_sheet_obj = wb_obj.active
    my_cell_obj = my_sheet_obj.cell(row = Row, column = column)
    return my_cell_obj.value
def read(path,header,rowkey):
    wb_obj = openpyxl.load_workbook(path)
    my_sheet_obj = wb_obj.active
    for i in range(1,my_sheet_obj.max_row+1):
        for j in range(1,my_sheet_obj.max_column+1):
            my_cell_obj = my_sheet_obj.cell(row = i, column = j)
            print(my_cell_obj.value)
    return my_cell_obj.value
print(getCoulums("myFile.xlsx"))  
print(getRows("myFile.xlsx"))   
print(readCellValue(1,1,"myFile.xlsx"))
print(read("myFile.xlsx","",""))